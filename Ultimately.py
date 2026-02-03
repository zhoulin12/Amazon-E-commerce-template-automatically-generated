import os
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def get_project_root():
    """获取项目根目录，兼容开发环境和PyInstaller打包环境"""
    try:
        # 检查是否有从main.py传递过来的project_root变量
        if 'project_root' in globals():
            return globals()['project_root']
        # PyInstaller创建一个临时文件夹并将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
        # 在打包环境中，可执行文件通常位于dist目录中
        # 我们需要获取可执行文件所在目录的上级目录作为项目根目录
        executable_dir = os.path.dirname(sys.executable)
        project_root = os.path.dirname(executable_dir)
    except Exception:
        # 检查是否有从main.py传递过来的project_root变量
        if 'project_root' in globals():
            return globals()['project_root']
        # 开发环境中，向上两级到达项目根目录
        # Release/Image Title Generation/Ultimately.py
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(os.path.dirname(script_dir))
    return project_root

def get_result_folder_from_config(config_file):
    """从配置文件中读取结果文件夹路径"""
    result_folder = None
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    if key.strip() == 'RESULT_FOLDER_PATH':
                        result_folder = value.strip()
                        break
    
    # 如果配置文件中没有指定结果文件夹路径，则使用默认路径
    if not result_folder:
        project_root = get_project_root()
        result_folder = os.path.join(project_root, "Result")
    
    # 确保结果文件夹存在
    os.makedirs(result_folder, exist_ok=True)
    return result_folder

def find_column_by_header(worksheet, header_text, row=4):
    """
    在指定行中查找包含特定文本的列
    
    Args:
        worksheet: Excel工作表对象
        header_text: 要查找的表头文本
        row: 查找的行号，默认为4
    
    Returns:
        int: 找到的列号，如果未找到返回None
    """
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=row, column=col).value
        if cell_value and header_text in str(cell_value):
            return col
    return None

def find_first_column_by_header(worksheet, header_text, row=4):
    """
    在指定行中查找第一个包含特定文本的列（用于有多个同名列的情况）
    
    Args:
        worksheet: Excel工作表对象
        header_text: 要查找的表头文本
        row: 查找的行号，默认为4
    
    Returns:
        int: 找到的第一个列号，如果未找到返回None
    """
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=row, column=col).value
        if cell_value and header_text in str(cell_value):
            return col
    return None

def find_column_by_multiple_headers(worksheet, header_texts, row=4):
    """
    在指定行中查找包含任一文本的列
    
    Args:
        worksheet: Excel工作表对象
        header_texts: 要查找的表头文本列表
        row: 查找的行号，默认为4
    
    Returns:
        int: 找到的列号，如果未找到返回None
    """
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=row, column=col).value
        if cell_value and any(header_text in str(cell_value) for header_text in header_texts):
            return col
    return None

def main():
    # 获取项目根目录
    project_root = get_project_root()
    print(f"Ultimately 脚本 - 项目根目录: {project_root}")
    
    # 从配置文件读取结果文件夹路径
    # config_file = Path(project_root) / "Release" / "config.txt"
    config_file = Path(project_root)  / "config.txt"
    result_dir = get_result_folder_from_config(config_file)
    
    # 输入文件和模板文件路径
    input_excel = Path(result_dir) / "Image_Titles_Add_Model.xlsx"
    
    # 查找模板文件，模糊匹配包含"上架模板"的文件
    template_dir = Path(project_root) / "需要的excel文件"
    template_file = None
    
    # 遍历目录中的所有Excel文件，查找包含"上架模板"的文件
    # 支持多种Excel文件扩展名：.xls, .xlsx, .xlsm
    excel_extensions = ["*.xls", "*.xlsx", "*.xlsm"]
    for extension in excel_extensions:
        for file_path in template_dir.glob(extension):
            if "上架模板" in file_path.name:
                template_file = file_path
                print(f"找到模板文件: {template_file.name}")
                break
        if template_file is not None:
            break
    
    # 如果没找到，抛出异常
    if template_file is None:
        # 收集所有Excel文件用于错误提示
        available_files = []
        for extension in excel_extensions:
            available_files.extend(template_dir.glob(extension))
        raise FileNotFoundError(f"在 {template_dir} 目录中未找到包含'上架模板'的文件。可用的文件: {[f.name for f in available_files]}")
    
    # 输出文件路径
    output_file = Path(result_dir) / "Final_Template.xlsm"
    
    # 检查输入文件是否存在
    if not input_excel.exists():
        print(f"输入文件不存在: {input_excel}")
        return
    
    # 检查模板文件是否存在
    if template_file is None or not template_file.exists():
        print(f"模板文件不存在: {template_file}")
        return
    
    print(f"正在读取输入文件: {input_excel}")
    print(f"正在读取模板文件: {template_file}")
    
    try:
        # 读取Excel文件
        df_input = pd.read_excel(input_excel)
        print(f"输入文件包含 {len(df_input)} 行数据")
        
        # 检查必要的列是否存在
        required_columns = ['图片名称', '亚马逊产品标题', '亚马逊产品标题翻译', '短标题', '短标题翻译', '图片编号', '型号']
        missing_columns = [col for col in required_columns if col not in df_input.columns]
        if missing_columns:
            print(f"输入文件缺少必要的列: {missing_columns}")
            return
        
        # 加载模板文件
        print("正在加载模板文件...")
        wb = load_workbook(template_file, keep_vba=True)  # 保持VBA宏
        
        # 尝试获取工作表，支持"模板"和"Template"两种名称
        template_sheet_names = ["模板", "Template"]
        ws = None
        for sheet_name in template_sheet_names:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"找到工作表: {sheet_name}")
                break
        
        # 如果都没找到，抛出异常
        if ws is None:
            raise ValueError(f"模板文件中未找到工作表: {template_sheet_names}。可用的工作表: {wb.sheetnames}")
        
        print(f"工作表最大行数: {ws.max_row}")
        print(f"工作表最大列数: {ws.max_column}")
        
        # 图片区域位置（根据调试信息，我们知道在第4行）
        image_start_row = 4  # 直接设置为第4行，如调试脚本确认的那样
        
        # 动态查找"主图像链接地址"列并确定图片链接列范围
        image_columns = []
        main_image_column = None
        
        # 查找"主图像链接地址"或"Main Image URL"列
        main_image_keywords = ["主图像链接地址", "Main Image URL"]
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=image_start_row, column=col).value
            if cell_value and any(keyword in str(cell_value) for keyword in main_image_keywords):
                main_image_column = col
                print(f"找到主图像列: {cell_value} (第{col}列)")
                break
        
        # 如果找到了"主图像链接地址"列，则确定7个图片链接列（主图像链接地址列及其后面6列）
        if main_image_column:
            image_columns = list(range(main_image_column, main_image_column + 7))
            print(f"动态找到图片链接列: {image_columns} (从第{main_image_column}列开始的7列)")
        else:
            # 如果没找到，使用默认的列范围（回退到原来的处理方式）
            print("警告: 未找到'主图像链接地址'或'Main Image URL'列，请检查列范围")
        
        # 动态查找其他需要替换的列
        product_name_col = find_column_by_multiple_headers(ws, ["产品名称", "产品名", "Product Name", "item_name"])  # 产品名称列
        
        # 定义需要填入“编号”（如 GYCYF000188）和“型号”（如 iPhone X）的列字母列表
        # 你可以直接在这里修改列字母，例如 ["CU", "B", "BC"]
        image_number_col_letters = ["BC", "BK"]  #编号
        model_col_letters = ["BE", "BX", "CU"]
        
        # 将列字母转换为索引列表
        image_number_columns = []
        for letter in image_number_col_letters:
            try:
                col_idx = column_index_from_string(letter)
                image_number_columns.append(col_idx)
            except Exception as e:
                print(f"警告: 无法解析编号列字母 '{letter}': {e}")
                
        model_columns = []
        for letter in model_col_letters:
            try:
                col_idx = column_index_from_string(letter)
                model_columns.append(col_idx)
            except Exception as e:
                print(f"警告: 无法解析型号列字母 '{letter}': {e}")
        
        # 检查是否找到所有必需的列
        if not product_name_col:
            print("警告: 未找到'产品名称'列，默认使用第4列")
            product_name_col = 4
        
        print(f"动态找到的列 - 产品名称: {product_name_col}")
        print(f"指定的编号列 (字母: {image_number_col_letters}) -> 索引: {image_number_columns}")
        print(f"指定的型号列 (字母: {model_col_letters}) -> 索引: {model_columns}")
        
        print(f"图片链接列: {image_columns}")
        
        # SKU列
        sku_col = 1  # A列
        # product_name_col 已经在上方动态确定了
        
        # 创建SKU映射列表（按顺序）
        sku_mapping_list = []  # 按顺序存储新SKU
        product_name_mapping_list = []  # 按顺序存储新产品名称
        image_number_mapping_list = []  # 按顺序存储图片编号
        model_type_mapping_list = []  # 按顺序存储型号
        
        # 填充映射列表
        for i in range(len(df_input)):
            new_sku = str(df_input.iloc[i]['图片名称'])
            new_product_name = str(df_input.iloc[i]['亚马逊产品标题'])
            image_number = str(df_input.iloc[i]['图片编号'])
            model_type = str(df_input.iloc[i]['型号'])
            
            sku_mapping_list.append(new_sku)
            product_name_mapping_list.append(new_product_name)
            image_number_mapping_list.append(image_number)
            model_type_mapping_list.append(model_type)
        
        print(f"映射关系已建立，总计 {len(sku_mapping_list)} 个映射")
        
        # 显示前5个映射
        print("前5个映射:")
        for i in range(min(5, len(sku_mapping_list))):
            print(f"  映射 {i}: {sku_mapping_list[i][:30]}...")
        
        # 步骤1: 替换SKU和产品名称（替换整列）
        print("\n开始替换SKU和产品名称...")
        
        # 数据从第8行开始（如用户要求）
        data_start_row = 8
        print(f"数据从第 {data_start_row} 行开始")
        
        # 收集所有有效的SKU行
        valid_rows = []  # 存储 (行号, 原始SKU) 元组
        row_idx = data_start_row
        
        while row_idx <= ws.max_row:
            # 检查当前行是否有SKU数据
            sku_cell = ws.cell(row=row_idx, column=sku_col).value
            if sku_cell and str(sku_cell).strip():
                original_sku = str(sku_cell)
                valid_rows.append((row_idx, original_sku))
            row_idx += 1
        
        print(f"在模板中找到 {len(valid_rows)} 个有效的数据行")
        
        # 检查是否需要添加更多行以容纳所有数据
        if len(df_input) > len(valid_rows):
            rows_needed = len(df_input) - len(valid_rows)
            print(f"需要添加 {rows_needed} 行以容纳所有数据")
            
            # 在末尾添加新行
            for i in range(rows_needed):
                # 从最后一行复制格式
                last_row_idx = valid_rows[-1][0] if valid_rows else data_start_row
                new_row_idx = ws.max_row + 1
                
                # 从最后一行复制行格式和公式
                for col in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row=last_row_idx, column=col)
                    target_cell = ws.cell(row=new_row_idx, column=col)
                    
                    # 复制值（但我们稍后会覆盖SKU和产品名称）
                    target_cell.value = source_cell.value
                    
                    # 复制格式
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()
                
                # 添加到valid_rows，使用空SKU（稍后会填充）
                valid_rows.append((new_row_idx, ""))
            
            print(f"已添加 {rows_needed} 个新行")
        
        # 遍历有效行进行替换
        sku_replacement_count = 0
        product_name_replacement_count = 0
        
        # 记录要保留的行号
        rows_to_keep = set()
        
        for i, (row_idx, original_sku) in enumerate(valid_rows):
            # 检查是否有对应的映射
            if i < len(sku_mapping_list):
                # 替换SKU
                new_sku = sku_mapping_list[i]
                ws.cell(row=row_idx, column=sku_col).value = new_sku
                sku_replacement_count += 1
                
                # 添加到保留集合
                rows_to_keep.add(row_idx)
                
                # 显示前5个替换
                if sku_replacement_count <= 5:
                    print(f"  替换SKU (第 {row_idx} 行): {original_sku[:30]}... -> {new_sku[:30]}...")
                
                # 替换产品名称
                new_product_name = product_name_mapping_list[i]
                ws.cell(row=row_idx, column=product_name_col).value = new_product_name
                product_name_replacement_count += 1
                
                # 显示前5个替换
                if product_name_replacement_count <= 5:
                    print(f"  替换产品名称 (第 {row_idx} 行): {new_product_name[:50]}...")
            
            # 显示进度
            if (i + 1) % 100 == 0:
                print(f"  已处理 {i + 1} 行...")
        
        print(f"总计替换了 {sku_replacement_count} 个SKU")
        print(f"总计替换了 {product_name_replacement_count} 个产品名称")
        
        # 步骤2: 更新图片链接中的SKU（替换整列）
        print("\n开始更新图片链接...")
        
        # 遍历有效行进行图片链接更新
        processed_rows = 0
        updated_links_count = 0
        
        for i, (row_idx, original_sku) in enumerate(valid_rows):
            # 检查是否有对应的映射
            if i < len(sku_mapping_list):
                new_sku = sku_mapping_list[i]
                
                # 更新图片链接列（只处理找到的7个图片链接列）
                row_updated_links = 0
                for col_idx, col in enumerate(image_columns):
                    link_cell = ws.cell(row=row_idx, column=col)
                    
                    # 从模板第8行读取对应列的图片链接格式
                    template_url = ws.cell(row=data_start_row, column=col).value
                    
                    # 如果模板中有URL，则基于模板URL构建新的URL；否则使用默认格式
                    if template_url and isinstance(template_url, str):
                        # 提取URL的各个部分
                        # 例如: http://geyishuma.com/GYFGCX0031GYFGCX0060/GYFGCX0031iPhone12.MAIN.jpg
                        import re
                        match = re.search(r'(http://[^/]+/)([^/]+/)([^.]+)(\..+)$', template_url)
                        if match:
                            base_url = match.group(1)  # http://geyishuma.com/
                            path_prefix = match.group(2)  # GYFGCX0031GYFGCX0060/
                            old_sku_part = match.group(3)  # GYFGCX0031iPhone12
                            suffix = match.group(4)  # .MAIN.jpg 或 .PT01.jpg 等
                            
                            # 根据列索引确定后缀
                            if col_idx == 0:
                                # 第一列（主图像链接地址）使用.MAIN
                                new_suffix = ".MAIN.jpg"
                            else:
                                # 其他列（其他图片链接地址）使用.PT加两位数字格式
                                new_suffix = f".PT{col_idx:02d}.jpg"
                            
                            # 构造新的图片链接URL
                            image_url = f"{base_url}{path_prefix}{new_sku}{new_suffix}"
                        else:
                            # 如果无法解析URL格式，则使用默认格式
                            if col_idx == 0:
                                image_url = f"http://geyishuma.com/GYFGCX0031GYFGCX0060/{new_sku}.MAIN.jpg"
                            else:
                                image_url = f"http://geyishuma.com/GYFGCX0031GYFGCX0060/{new_sku}.PT{col_idx:02d}.jpg"
                    else:
                        # 如果模板中没有URL或不是字符串，则使用默认格式
                        if col_idx == 0:
                            image_url = f"http://geyishuma.com/GYFGCX0031GYFGCX0060/{new_sku}.MAIN.jpg"
                        else:
                            image_url = f"http://geyishuma.com/GYFGCX0031GYFGCX0060/{new_sku}.PT{col_idx:02d}.jpg"
                    
                    # 设置图片链接
                    link_cell.value = image_url
                    row_updated_links += 1
                    updated_links_count += 1
                
                if row_updated_links > 0 and processed_rows < 5:  # 显示前5个更新
                    print(f"  更新图片链接 (第 {row_idx} 行) {original_sku[:30]}... -> {new_sku[:30]}... ({row_updated_links} 个链接)")
                
                processed_rows += 1
            
            # 显示进度
            if (i + 1) % 100 == 0:
                print(f"  已处理 {i + 1} 行...")
        
        print(f"总计处理了 {processed_rows} 行，更新了 {updated_links_count} 个图片链接")
        
        # 步骤3: 替换编号列内容（如零件编号）
        print("\n开始替换编号相关列...")
        
        data_start_row = 8
        image_number_replaced_count = 0
        
        if image_number_columns:
            for i, (row_idx, original_sku) in enumerate(valid_rows):
                if i < len(image_number_mapping_list) and row_idx >= data_start_row:
                    new_image_number = image_number_mapping_list[i]
                    for col in image_number_columns:
                        ws.cell(row=row_idx, column=col).value = new_image_number
                    image_number_replaced_count += 1
            print(f"总计替换了 {len(image_number_columns)} 个列中的 {image_number_replaced_count} 行编号数据")
        else:
            print("未找到任何编号列，跳过替换")
        
        # 步骤4: 替换型号相关列内容（如型号、兼容设备等）
        print("\n开始替换型号相关列...")
        
        model_replaced_count = 0
        
        if model_columns:
            for i, (row_idx, original_sku) in enumerate(valid_rows):
                if i < len(model_type_mapping_list) and row_idx >= data_start_row:
                    new_model_type = model_type_mapping_list[i]
                    for col in model_columns:
                        ws.cell(row=row_idx, column=col).value = new_model_type
                    model_replaced_count += 1
            print(f"总计替换了 {len(model_columns)} 个列中的 {model_replaced_count} 行型号数据")
        else:
            print("未找到任何型号列，跳过替换")
        
        # 步骤7: 删除多余行
        print("\n开始删除多余行...")
        rows_deleted = 0
        
        # 从最后一行向上删除以避免行号变化
        row_idx = ws.max_row
        while row_idx >= data_start_row:
            if row_idx not in rows_to_keep:
                ws.delete_rows(row_idx)
                rows_deleted += 1
            row_idx -= 1
        
        print(f"删除了 {rows_deleted} 个多余行")
        
        # 步骤8: 统计产品名称列字符数并在最后一列显示
        print(f"\n开始统计第 {product_name_col} 列字符数...")
        
        # 获取最后一列的索引
        last_col = ws.max_column + 1  # 在最后一列之后添加新列
        
        # 统计字符数并写入最后一列
        char_count_added = 0
        for i, (row_idx, original_sku) in enumerate(valid_rows):
            if row_idx >= data_start_row:
                # 获取产品名称列的值
                name_cell_value = ws.cell(row=row_idx, column=product_name_col).value
                if name_cell_value is not None:
                    # 计算字符数
                    char_count = len(str(name_cell_value))
                    # 在最后一列写入字符数
                    ws.cell(row=row_idx, column=last_col).value = char_count
                    char_count_added += 1
                    
                    # 显示前5个统计
                    if char_count_added <= 5:
                        print(f"  第 {row_idx} 行产品名称字符数: {char_count}")
        
        print(f"总计为 {char_count_added} 行添加了字符数统计")
        
        # 检查文件是否存在，如果存在则重命名
        final_output_file = output_file
        counter = 1
        while final_output_file.exists():
            # 获取文件名和扩展名
            name, ext = final_output_file.stem, final_output_file.suffix
            # 在文件名中添加计数器
            final_output_file = Path(final_output_file.parent) / f"{name}_{counter}{ext}"
            counter += 1
    
        # 自动插入父类模板行
        print("\n开始自动插入父类模板行...")
        try:
            # 查找"父条目的库存单位"列
            parent_sku_col = None
            for cell in ws[4]:  # 第4行是标题行
                if cell.value and "父条目的库存单位" in str(cell.value):
                    parent_sku_col = cell.column
                    break
            
            if parent_sku_col:
                print(f"找到'父条目的库存单位'列，位于第 {parent_sku_col} 列")
                
                # 读取Excel文件数据来获取父类编号信息
                # 使用Add Model.py生成的文件
                project_root = get_project_root()
                # 从配置文件中读取结果文件夹路径
                config_file = Path(project_root) / "config.txt"
                result_dir_path = "Result"  # 默认路径
                if config_file.exists():
                    with open(config_file, 'r', encoding='utf-8') as f:
                        for line in f:
                            line = line.strip()
                            if line and not line.startswith('#') and '=' in line:
                                key, value = line.split('=', 1)
                                if key.strip() == 'RESULT_FOLDER_PATH':
                                    result_dir_path = value.strip()
                                    break
                
                result_dir = Path(result_dir_path)
                input_excel = result_dir / "Image_Titles_Add_Model.xlsx"
                print(f"使用Add Model生成的输入文件: {input_excel}")
                
                if input_excel.exists():
                    print(f"正在读取输入文件: {input_excel}")
                    df_input = pd.read_excel(input_excel)
                    
                    # 检查是否有父类编号列
                    if '父类编号' in df_input.columns:
                        # 按父类编号分组统计数据
                        parent_stats = df_input.groupby('父类编号').size().reset_index(name='count')
                        print(f"找到 {len(parent_stats)} 个父类编号")
                        
                        # 收集所有需要处理的SKU及其对应的父类编号
                        sku_to_parent = {}
                        for idx, row in df_input.iterrows():
                            sku = row['图片名称']
                            parent_id = row['父类编号']
                            sku_to_parent[sku] = parent_id
                        
                        # 收集所有唯一的父类编号，按出现顺序排列
                        unique_parents = []
                        seen_parents = set()
                        for sku in sku_to_parent:
                            parent_id = sku_to_parent[sku]
                            if parent_id not in seen_parents:
                                unique_parents.append(parent_id)
                                seen_parents.add(parent_id)
                        
                        print(f"需要处理的父类编号顺序: {unique_parents}")
                        
                        # 在每个父类的第一行插入父类模板行
                        inserted_rows = 0
                        processed_parents = set()
                        
                        # 收集所有需要插入父类行的位置
                        insert_positions = []  # [(row_position, parent_id), ...]
                        
                        # 从第8行开始遍历数据行，收集插入位置
                        row_idx = 8
                        max_row = ws.max_row
                        
                        while row_idx <= max_row:
                            # 获取SKU值来查找对应的父类编号
                            sku_cell = ws.cell(row=row_idx, column=1)  # SKU在A列
                            if sku_cell.value:
                                sku_value = str(sku_cell.value)
                                # 在输入数据中查找对应的父类编号
                                if sku_value in sku_to_parent:
                                    parent_id = sku_to_parent[sku_value]
                                    
                                    # 如果这是该父类的第一行且尚未处理过
                                    if parent_id not in processed_parents:
                                        processed_parents.add(parent_id)
                                        insert_positions.append((row_idx + inserted_rows, parent_id))
                                        print(f"  计划在第 {row_idx + inserted_rows} 行插入父类模板行，父类编号: {parent_id}")
                            
                            row_idx += 1
                        
                        # 按照行号从高到低排序，这样插入不会影响后续行号
                        insert_positions.sort(reverse=True)
                        
                        # 执行插入操作
                        for row_pos, parent_id in insert_positions:
                            # 插入新行
                            ws.insert_rows(row_pos)
                            inserted_rows += 1
                            
                            # 复制第7行的所有内容（参考行）
                            reference_row = 7
                            for col in range(1, ws.max_column + 1):
                                source_cell = ws.cell(row=reference_row, column=col)
                                target_cell = ws.cell(row=row_pos, column=col)
                                
                                # 复制值
                                target_cell.value = source_cell.value
                                
                                # 复制格式
                                target_cell.font = source_cell.font.copy()
                                target_cell.border = source_cell.border.copy()
                                target_cell.fill = source_cell.fill.copy()
                                target_cell.number_format = source_cell.number_format
                                target_cell.protection = source_cell.protection.copy()
                                target_cell.alignment = source_cell.alignment.copy()
                            
                            # 修改特定列的内容
                            # A列: 父类编号（而不是"父条目"）
                            ws.cell(row=row_pos, column=1).value = parent_id
                            
                            # 产品名称列: 父类SKU（保持不变，已经是父类编号）
                            ws.cell(row=row_pos, column=product_name_col).value = parent_id
                            
                            # "父条目的库存单位"列: 父类SKU
                            ws.cell(row=row_pos, column=parent_sku_col).value = parent_id
                            
                            # 清除所有图片链接列，父类行不应包含图片链接
                            for col in image_columns:
                                ws.cell(row=row_pos, column=col).value = None
                            
                            print(f"  在第 {row_pos} 行插入父类模板行，父类编号: {parent_id}")
                        
                        # 更新所有子类行的"父条目的库存单位"列为对应的父类编号
                        print("开始更新子类行的'父条目的库存单位'列...")
                        updated_subclass_count = 0
                        
                        # 重新遍历所有行（包括新插入的父类行）
                        row_idx = 8
                        max_row = ws.max_row
                        
                        while row_idx <= max_row:
                            # 获取SKU值
                            sku_cell = ws.cell(row=row_idx, column=1)  # SKU在A列
                            if sku_cell.value:
                                sku_value = str(sku_cell.value)
                                # 检查是否是子类SKU（在输入数据中可以找到）
                                if sku_value in sku_to_parent:
                                    parent_id = sku_to_parent[sku_value]
                                    # 更新"父条目的库存单位"列
                                    ws.cell(row=row_idx, column=parent_sku_col).value = parent_id
                                    updated_subclass_count += 1
                                    
                                    # 显示前几个更新
                                    if updated_subclass_count <= 5:
                                        print(f"  更新第 {row_idx} 行的'父条目的库存单位'列: {parent_id}")
                            
                            row_idx += 1
                        
                        print(f"总计更新了 {updated_subclass_count} 个子类行的'父条目的库存单位'列")
                        print(f"总计插入了 {inserted_rows} 个父类模板行")
                    else:
                        print("输入文件中未找到'父类编号'列")
                else:
                    print(f"输入文件不存在: {input_excel}")
            else:
                print("未找到'父条目的库存单位'列，跳过父类模板行插入")
        except Exception as e:
            print(f"插入父类模板行时出错: {e}")
            import traceback
            traceback.print_exc()
    
        # 保存文件
        print(f"\n正在保存处理后的文件到: {final_output_file}")
        try:
            wb.save(final_output_file)
            print("处理完成!")
        except Exception as e:
            raise Exception(f"保存文件时出错: {e}\n请确保Excel文件未被其他程序打开")
        
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()