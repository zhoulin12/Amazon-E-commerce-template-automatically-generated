import os
import sys
from pathlib import Path
from openpyxl import load_workbook
import shutil
import pandas as pd

def get_project_root():
    """获取项目根目录，兼容开发环境和PyInstaller打包环境"""
    try:
        # 检查是否有从main.py传递过来的project_root变量
        if 'project_root' in globals():
            return globals()['project_root']
        # PyInstaller创建一个临时文件夹并将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
        # 在打包环境中，可执行文件通常位于dist目录中
        # 我们需要获取可执行文件所在目录的上级目录作为项目根目录
        executable_dir = os.path.dirname(sys.executable)
        project_root = os.path.dirname(executable_dir)
    except Exception:
        # 检查是否有从main.py传递过来的project_root变量
        if 'project_root' in globals():
            return globals()['project_root']
        # 开发环境中，向上两级到达项目根目录
        # Release/Image Title Generation/Organize.py
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(os.path.dirname(script_dir))
    return project_root

def get_result_folder_from_config(config_file):
    """从配置文件中读取结果文件夹路径"""
    result_folder = None
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    if key.strip() == 'RESULT_FOLDER_PATH':
                        result_folder = value.strip()
                        break
    
    # 如果配置文件中没有指定结果文件夹路径，则使用默认路径
    if not result_folder:
        project_root = get_project_root()
        result_folder = os.path.join(project_root, "Result")
    
    # 确保结果文件夹存在
    os.makedirs(result_folder, exist_ok=True)
    return result_folder

def split_excel_by_phone_model():
    """
    对"模板"工作表中的数据进行整理：
    1. 复制原始文件生成两个新文件（保持所有格式和其他工作表）
    2. 在iPhone文件中只保留iPhone数据行
    3. 在Samsung文件中只保留Samsung数据行
    """
    
    # 获取项目根目录
    project_root = get_project_root()
    print(f"Organize 脚本 - 项目根目录: {project_root}")
    
    # 从配置文件读取结果文件夹路径
    config_file = Path(project_root) / "config.txt"
    result_dir = get_result_folder_from_config(config_file)
    
    # 输入文件路径
    input_file = Path(result_dir) / "Final_Template.xlsm"
    
    # 检查输入文件是否存在
    if not input_file.exists():
        print(f"输入文件不存在: {input_file}")
        return
    
    print(f"正在读取文件: {input_file}")
    
    try:
        # 加载工作簿并选择"模板"工作表
        wb = load_workbook(input_file, keep_vba=True)
        ws = wb["模板"]
        
        # 获取所有行数据
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"工作表共有 {max_row} 行, {max_col} 列")
        
        # 读取所有行数据（从第8行开始是数据行）
        data_start_row = 8
        data_rows = []
        
        # 读取数据行（从第8行开始）
        for row_idx in range(data_start_row, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            data_rows.append(row_data)
        
        # 读取父类信息
        parent_info = {}  # 存储SKU到父类编号的映射
        try:
            # 从Image_Titles_Add_Model.xlsx读取父类信息
            input_excel = Path(result_dir) / "Image_Titles_Add_Model.xlsx"
            if input_excel.exists():
                df_input = pd.read_excel(input_excel)
                if '父类编号' in df_input.columns and '图片名称' in df_input.columns:
                    for idx, row in df_input.iterrows():
                        sku = row['图片名称']
                        parent_id = row['父类编号']
                        parent_info[sku] = parent_id
        except Exception as e:
            print(f"读取父类信息时出错: {e}")
        
        # 分离iPhone和samsung行，同时正确排列父类行和子类行
        def organize_rows_by_phone_model(phone_type):
            """根据手机类型分离行并正确排列父类行和子类行"""
            organized_rows = []
            processed_parents = set()  # 已处理的父类
            
            # 遍历所有数据行
            for row in data_rows:
                if row and row[0]:  # 确保A列有值
                    cell_value = str(row[0])
                    # 检查是否符合指定的手机类型
                    if (phone_type == 'iPhone' and 'iPhone' in cell_value) or \
                       (phone_type == 'Samsung' and 'samsung' in cell_value.lower()):
                        # 检查是否有对应的父类信息
                        if cell_value in parent_info:
                            parent_id = parent_info[cell_value]
                            # 如果这是该父类的第一行且尚未处理过，则先添加父类行
                            if parent_id not in processed_parents:
                                processed_parents.add(parent_id)
                                # 查找并添加父类行
                                for parent_row in data_rows:
                                    if parent_row and parent_row[0] and str(parent_row[0]) == parent_id:
                                        organized_rows.append(parent_row)
                                        break
                        # 添加当前行
                        organized_rows.append(row)
            return organized_rows
        
        # 为iPhone和Samsung行正确排列父类行和子类行
        iphone_rows = organize_rows_by_phone_model('iPhone')
        samsung_rows = organize_rows_by_phone_model('Samsung')
        
        # 其他行保持不变
        other_rows = []
        for row in data_rows:
            if row and row[0]:  # 确保A列有值
                cell_value = str(row[0])
                if 'iPhone' not in cell_value and 'samsung' not in cell_value.lower():
                    other_rows.append(row)
            else:
                # 如果A列没有值，也归类到other_rows
                other_rows.append(row)
        
        print(f"找到 {len(iphone_rows)} 行包含'iPhone'")
        print(f"找到 {len(samsung_rows)} 行包含'samsung'")
        print(f"找到 {len(other_rows)} 行包含其他内容")
        
        # 处理iPhone文件
        if iphone_rows:
            # 复制原始文件为iPhone版本
            iphone_output_file = Path(result_dir) / "Final_Template_iPhone.xlsm"
            counter = 1
            while iphone_output_file.exists():
                name, ext = iphone_output_file.stem, iphone_output_file.suffix
                iphone_output_file = Path(iphone_output_file.parent) / f"{name}_{counter}{ext}"
                counter += 1
            
            # 复制文件
            shutil.copy2(input_file, iphone_output_file)
            
            # 打开复制的文件并修改"模板"工作表
            iphone_wb = load_workbook(iphone_output_file, keep_vba=True)
            iphone_ws = iphone_wb["模板"]
            
            # 删除原来的数据行（从第8行开始）
            for row_idx in range(max_row, data_start_row - 1, -1):
                iphone_ws.delete_rows(row_idx)
            
            # 写入iPhone数据行
            current_row = data_start_row
            for row in iphone_rows:
                for col_idx, value in enumerate(row, 1):
                    iphone_ws.cell(row=current_row, column=col_idx).value = value
                current_row += 1
            
            # 查找"父条目的库存单位"列并添加前缀"P"
            parent_sku_col = None
            for cell in iphone_ws[4]:  # 第4行是标题行
                if cell.value and "父条目的库存单位" in str(cell.value):
                    parent_sku_col = cell.column
                    break
            
            if parent_sku_col:
                # 从第8行开始处理数据行
                for row_idx in range(8, iphone_ws.max_row + 1):
                    cell_value = iphone_ws.cell(row=row_idx, column=parent_sku_col).value
                    if cell_value:
                        # 添加前缀"P"
                        iphone_ws.cell(row=row_idx, column=parent_sku_col).value = "P-" + str(cell_value)
            
            # 保存iPhone文件
            iphone_wb.save(iphone_output_file)
            print(f"iPhone数据已保存到: {iphone_output_file}")
        
        # 处理Samsung文件
        if samsung_rows:
            # 复制原始文件为Samsung版本
            samsung_output_file = Path(result_dir) / "Final_Template_Samsung.xlsm"
            counter = 1
            while samsung_output_file.exists():
                name, ext = samsung_output_file.stem, samsung_output_file.suffix
                samsung_output_file = Path(samsung_output_file.parent) / f"{name}_{counter}{ext}"
                counter += 1
            
            # 复制文件
            shutil.copy2(input_file, samsung_output_file)
            
            # 打开复制的文件并修改"模板"工作表
            samsung_wb = load_workbook(samsung_output_file, keep_vba=True)
            samsung_ws = samsung_wb["模板"]
            
            # 删除原来的数据行（从第8行开始）
            for row_idx in range(max_row, data_start_row - 1, -1):
                samsung_ws.delete_rows(row_idx)
            
            # 写入Samsung数据行
            current_row = data_start_row
            for row in samsung_rows:
                for col_idx, value in enumerate(row, 1):
                    samsung_ws.cell(row=current_row, column=col_idx).value = value
                current_row += 1
            
            # 查找"父条目的库存单位"列并添加前缀"S"
            parent_sku_col = None
            for cell in samsung_ws[4]:  # 第4行是标题行
                if cell.value and "父条目的库存单位" in str(cell.value):
                    parent_sku_col = cell.column
                    break
            
            if parent_sku_col:
                # 从第8行开始处理数据行
                for row_idx in range(8, samsung_ws.max_row + 1):
                    cell_value = samsung_ws.cell(row=row_idx, column=parent_sku_col).value
                    if cell_value:
                        # 添加前缀"S"
                        samsung_ws.cell(row=row_idx, column=parent_sku_col).value = "S-" + str(cell_value)
            
            # 保存Samsung文件
            samsung_wb.save(samsung_output_file)
            print(f"Samsung数据已保存到: {samsung_output_file}")
        
        print("文件拆分完成！")
        
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        import traceback
        traceback.print_exc()

def main():
    """主函数"""
    split_excel_by_phone_model()

if __name__ == "__main__":
    main()